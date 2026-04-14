import streamlit as st
import pandas as pd
import json
import io
import os
import zipfile
import shutil
import openpyxl
from datetime import datetime

# ======================== 【双密码独立配置】 ========================
LOGIN_PASSWORD = "123123"       # 普通用户登录密码（自行修改）
ADMIN_PASSWORD = "888888"     # 管理员密码（自行修改）
# ==================================================================

# 系统登录校验
def check_system_login():
    if "is_logged_in" not in st.session_state:
        st.session_state.is_logged_in = False
    if not st.session_state.is_logged_in:
        st.title("🔒 系统登录")
        pwd = st.text_input("请输入登录密码", type="password")
        if st.button("登录系统"):
            if pwd == LOGIN_PASSWORD:
                st.session_state.is_logged_in = True
                st.success("登录成功！")
                st.rerun()
            else:
                st.error("登录密码错误！")
        st.stop()

check_system_login()

# ===========================
# 🔥 永久存储路径（GitHub 仓库固定路径，永不丢失）
# ===========================
# 模板/映射直接读取 GitHub 仓库文件，重启不丢失
TEMPLATE_FOLDER = "b_templates"
MAPPING_FOLDER = "template_mappings"

# ===========================
# 文本读取（防ID尾数变0 + 超强容错）
# ===========================
def read_excel(file, sheet_name=0):
    """
    读取Excel文件
    - sheet_name=0（默认）：读取第一个工作表，适用于A表、C表和B模板
    """
    try:
        df = pd.read_excel(
            file, 
            engine="openpyxl", 
            dtype=str, 
            header=0,       # 强制指定：第一行=列名
            sheet_name=sheet_name,
            usecols=None    # 自动识别有效列
        )
        # 自动删除所有 Unnamed 无名列
        df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
        # 清理列名的隐形空格/乱码
        df.columns = df.columns.str.strip()
        return df
    except Exception as e:
        st.error(f"文件读取失败：{str(e)}")
        return None

# ===========================
# 🔥 从 GitHub 仓库读取模板（永久存储）
# ===========================
def get_b_templates(a_type):
    if not os.path.exists(TEMPLATE_FOLDER):
        return []
    all_files = [f for f in os.listdir(TEMPLATE_FOLDER) if f.endswith((".xlsx", ".xls"))]
    if a_type == "旧A":
        # 旧A：返回所有模板（旧的没有_a_type标记，混在一起）
        return all_files
    # 新A：只返回 _a_type == "新A" 的模板
    result = []
    for f in all_files:
        mp = load_mapping(f, None, None, skip_a_check=True)
        if mp.get("_a_type") == "新A":
            result.append(f)
    return result

# ===========================
# 自动映射 + 缺失列检测 + 永久映射管理
# ===========================
def auto_map_columns(df_b, df_a):
    auto_mapping = {}
    b_cols = list(df_b.columns)
    a_cols = list(df_a.columns)
    for col in b_cols:
        if col in a_cols:
            auto_mapping[col] = col
    return auto_mapping

def check_missing_columns(df_b, df_a):
    return [col for col in df_b.columns if col not in df_a.columns]

def load_mapping(template_name, df_b, df_a, skip_a_check=False):
    """skip_a_check=True 时返回完整mapping（含_a_type），不验证A表列"""
    map_file = os.path.join(MAPPING_FOLDER, f"{template_name}.json")
    if os.path.exists(map_file):
        with open(map_file, "r", encoding="utf-8") as f:
            mapping = json.load(f)
        if "_a_type" not in mapping:
            mapping["_a_type"] = "旧A"
        if skip_a_check:
            return mapping
        return {k: v for k, v in mapping.items() if k != "_a_type"}
    if skip_a_check:
        return {"_a_type": "旧A"}
    return auto_map_columns(df_b, df_a)

def save_mapping(template_name, mapping, a_type):
    os.makedirs(MAPPING_FOLDER, exist_ok=True)
    map_file = os.path.join(MAPPING_FOLDER, f"{template_name}.json")
    full_mapping = {"_a_type": a_type}
    full_mapping.update(mapping)
    with open(map_file, "w", encoding="utf-8") as f:
        json.dump(full_mapping, f, ensure_ascii=False, indent=2)

# ===========================
# 三表匹配引擎（终极容错）
# ===========================
class DataMatcher:
    def __init__(self, df_a):
        self.df_a = df_a.copy().astype(str).apply(lambda x: x.str.strip())

    def match_c_to_a(self, df_c, key_c, key_a):
        df_c_clean = df_c.copy().astype(str).apply(lambda x: x.str.strip())
        return pd.merge(df_c_clean[[key_c]], self.df_a, left_on=key_c, right_on=key_a, how="left")

    def fill_b_template(self, df_matched, df_b_template, mapping):
        df_b = df_b_template.copy().astype(str).apply(lambda x: x.str.strip())
        for b_col, a_col in mapping.items():
            if b_col in df_b.columns and a_col in df_matched.columns:
                # 清理 NaN 和 'nan' 字符串
                cleaned_values = df_matched[a_col].apply(
                    lambda x: None if pd.isna(x) or (isinstance(x, str) and x.lower() == 'nan') else x
                )
                df_b[b_col] = cleaned_values
        return df_b

# ===========================
# 🔥 基于模板生成结果（保留附表，只改 Sheet1）
# ===========================
def convert_to_native(val):
    """
    将字符串值转换为原生类型，保证Excel格式正确
    身份证等长数字保持为文本，防止科学计数法
    """
    # 处理 NaN 和 'nan' 字符串
    if pd.isna(val) or (isinstance(val, str) and val.lower() == 'nan'):
        return None
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    if isinstance(val, str):
        val = val.strip()
        # 身份证/编号等长数字必须保持为文本：
        # 1. 长度 > 15 位
        # 2. 首位为 0
        # 3. 包含非数字字符
        if len(val) > 15 or val.startswith('0'):
            return val
        
        # 尝试转换为数字
        try:
            if "." in val:
                return float(val)
            else:
                num_val = int(val)
                # 如果数字太小（比如序号1,2,3），当作数字
                # 但身份证位数的大数字，返回文本
                return val if num_val > 10**15 else num_val
        except ValueError:
            return val
    return val

def generate_result_from_template(template_path, result_df, output_sheet_name=None):
    """
    基于B模板文件生成结果Excel：
    - 复制整个模板文件（保留所有附表、格式、公式、图片等）
    - 只替换第一个工作表的数据行（Sheet1或第一个sheet）
    - 返回 BytesIO 对象
    """
    if output_sheet_name is None:
        # 默认使用模板的第一个sheet
        wb_temp = openpyxl.load_workbook(template_path)
        output_sheet_name = wb_temp.sheetnames[0]
        wb_temp.close()
    buf = io.BytesIO()
    # 把模板文件完整复制到内存
    with open(template_path, "rb") as f:
        buf.write(f.read())
    buf.seek(0)

    # 用 openpyxl 打开副本，只操作目标 sheet
    wb = openpyxl.load_workbook(buf)

    # 如果目标 sheet 不存在，直接写入（兜底）
    if output_sheet_name in wb.sheetnames:
        ws = wb[output_sheet_name]
        # 从第2行开始写数据（第1行保留模板表头）
        for r_idx, row in enumerate(result_df.values, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=convert_to_native(val))
    else:
        ws = wb.create_sheet(output_sheet_name)
        # 写表头
        for c_idx, col_name in enumerate(result_df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)
        # 写数据
        for r_idx, row in enumerate(result_df.values, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=convert_to_native(val))

    # 保存到 BytesIO
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    wb.close()
    return out

# ===========================
# 批量生成打包 ZIP
# ===========================
def create_zip(files_data):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, data in files_data.items():
            zf.writestr(filename, data)
    zip_buffer.seek(0)
    return zip_buffer

# ===========================
# 界面布局
# ===========================
st.set_page_config(page_title="三表Excel生成器", layout="wide")
st.title("📊 一键智能vlookup")

# ===========================
# 侧边栏（A表类型切换 + 数据上传）
# ===========================
with st.sidebar:
    st.header("选择表类型")
    # a_type_val 统一用 "旧A" / "新A"，避免与 radio option 字符串混淆
    if st.session_state.get("a_type_val") is None:
        st.session_state.a_type_val = "旧A"
    idx = 0 if st.session_state.a_type_val == "旧A" else 1
    a_type_display = st.radio(
        "表类型",
        ["上传表", "要素表"],
        index=idx,
        key="a_type_radio",
        horizontal=True,
        help="切换后上传区和模板列表会自动切换"
    )
    # 同步到 a_type_val
    a_type_val = "旧A" if a_type_display == "上传表" else "新A"
    if st.session_state.a_type_val != a_type_val:
        st.session_state.a_type_val = a_type_val
        # 切换A表类型时清空状态
        for key in list(st.session_state.keys()):
            if key not in ["is_logged_in", "a_type_val", "a_type_radio"]:
                try:
                    del st.session_state[key]
                except:
                    pass
        st.rerun()

    st.divider()
    st.header("1. 数据上传")
    up_a = st.file_uploader("A表（总数据源）", type=['xlsx', 'xls'])
    up_c = st.file_uploader("C表（只需主键列）", type=['xlsx', 'xls'])

# ===========================
# 主功能区
# ===========================
template_list = get_b_templates(st.session_state.get("a_type_val", "旧A"))

if up_a and up_c:
    df_a = read_excel(up_a)
    df_c = read_excel(up_c)
    if df_a is None or df_c is None:
        st.stop()

    # -------------------- 
    # 选择要生成的B模板（支持多选批量）
    # --------------------
    if not template_list:
        st.error("❌ 暂无可用模板，请联系管理员上传")
        st.stop()

    selected_templates = st.multiselect("选择要生成的B模板（可多选批量）", template_list)

    if selected_templates:
        # 预览上传的表
        col1, col2 = st.columns(2)
        with col1: st.subheader("A表（总数据源）"); st.dataframe(df_a.head(3), use_container_width=True)
        with col2: st.subheader("C表（主键）"); st.dataframe(df_c.head(3), use_container_width=True)

        # 主键选择
        st.subheader("⚙️ 主键匹配")
        key_c = st.selectbox("C表主键列", df_c.columns)
        key_a = st.selectbox("A表对应主键列", df_a.columns)

        # 对第一个模板展示映射配置界面（批量时所有模板各自用已保存的映射）
        first_template = selected_templates[0]
        template_path_preview = os.path.join(TEMPLATE_FOLDER, first_template)
        df_b_preview = read_excel(template_path_preview)

        start_gen = False
        current_map = {}

        if df_b_preview is not None:
            st.subheader("🔗 字段映射配置")
            st.caption(f"以下为【{first_template}】的映射，每个模板独立保存")
            saved_map = load_mapping(first_template, df_b_preview, df_a)

            map_cols = st.columns(3)
            for i, b_col in enumerate(df_b_preview.columns):
                opts = ['--- 不填充 ---'] + list(df_a.columns)
                default_val = saved_map.get(b_col, '--- 不填充 ---')
                default_idx = opts.index(default_val) if default_val in opts else 0
                with map_cols[i % 3]:
                    sel = st.selectbox(f"B→{b_col}", opts, index=default_idx, key=f"map_{first_template}_{b_col}")
                    if sel != '--- 不填充 ---':
                        current_map[b_col] = sel

            st.divider()
            col_save, col_gen = st.columns(2)
            with col_save:
                if st.button("💾 保存当前映射（永久生效）"):
                    save_mapping(first_template, current_map, st.session_state.get("a_type_val", "旧A"))
                    st.success("✅ 映射已永久保存！")
            with col_gen:
                start_gen = st.button("🚀 一键生成", type="primary")

        if start_gen:
            matcher = DataMatcher(df_a)
            bar = st.progress(0)
            status = st.empty()
            files_data = {}
            total = len(selected_templates)
            preview_result = None

            for idx, b_file in enumerate(selected_templates):
                progress = int((idx + 1) / total * 100)
                status.text(f"生成中：{b_file}")
                bar.progress(progress)

                t_path = os.path.join(TEMPLATE_FOLDER, b_file)
                df_b = read_excel(t_path)

                if df_b is None:
                    continue

                # 第一个模板用界面配置的映射，其余用各自已保存的映射
                if b_file == first_template:
                    mapping = current_map
                else:
                    mapping = load_mapping(b_file, df_b, df_a)

                matched = matcher.match_c_to_a(df_c, key_c, key_a)
                result = matcher.fill_b_template(matched, df_b, mapping)

                # 基于模板生成（保留附表）
                result_buf = generate_result_from_template(t_path, result)
                files_data[f"结果_{b_file}"] = result_buf.getvalue()

                # 保留第一个结果用于预览
                if preview_result is None:
                    preview_result = result

            bar.progress(100)
            status.text("✅ 全部生成完成！")

            if preview_result is not None:
                st.dataframe(preview_result, use_container_width=True)

            if len(files_data) == 1:
                # 单文件直接下载
                fname = list(files_data.keys())[0]
                st.download_button("📥 下载结果", files_data[fname], fname)
            else:
                # 多文件打包ZIP
                st.success(f"成功生成 {len(files_data)} 个文件")
                zip_file = create_zip(files_data)
                st.download_button(
                    "📦 下载全部文件(ZIP)",
                    zip_file,
                    f"批量生成结果_{datetime.now().strftime('%Y%m%d%H%M')}.zip"
                )

else:
    st.info("""
    ✅ 永久存储说明：
    1. 所有B模板 + 字段映射 均保存在GitHub，公网永不丢失
    2. 支持单模板生成 / 多模板批量一键生成
    3. 同名列自动映射，缺失列不报错
    4. 双密码权限管理
    """)